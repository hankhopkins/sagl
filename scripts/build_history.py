"""Extract SAGL history from 2023-2026 workbooks into history.json"""
import openpyxl, json, datetime, re, statistics
from collections import defaultdict

# 2026 is intentionally NOT here: the site fetches current-season data live
# from the Google Sheet and merges it client-side, so History/Profiles stay
# current without regenerating this file. Add 2026 after the season completes.
FILES = {
    2023: '/mnt/user-data/uploads/Stone_Arch_Golf_League_Data_2023_1_.xlsx',
    2024: '/mnt/user-data/uploads/Stone_Arch_Golf_League_Data_2024_1_.xlsx',
    2025: '/mnt/user-data/uploads/Stone_Arch_Golf_League_Data_2025_2_.xlsx',
}

def rows_of(ws):
    return [list(r) for r in ws.iter_rows(values_only=True)]

def clean(v):
    if v is None: return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v

NAME_FIX = {'Mark MIkhail': 'Mark Mikhail'}
JUNK_NAMES = {'Gross','Net','Match','Results','Name','Team 1','Team 2'}

def fix_name(v):
    if not v: return v
    v = NAME_FIX.get(v, v)
    return v

def is_team_row(name):
    return bool(name) and ' & ' in name

def is_junk(name):
    return (not name) or name in JUNK_NAMES or ',' in name

def num(v):
    try:
        f = float(v)
        return int(f) if f == int(f) else round(f, 2)
    except (TypeError, ValueError):
        return None

def datestr(v):
    if isinstance(v, datetime.datetime):
        return v.strftime('%-m/%-d/%Y')
    if isinstance(v, str) and v.strip():
        return v.strip()
    return None

def dateyear(v):
    if isinstance(v, datetime.datetime):
        return v.year
    return None

# ---------- Directory ----------
def parse_directory(wb):
    ws = wb['league_directory']
    rows = rows_of(ws)
    players = {}   # name -> team#
    for r in rows[1:]:
        team, name = num(r[0]), clean(r[1])
        if name and team:
            players[name] = int(team)
    teams = defaultdict(list)
    for name, t in players.items():
        teams[t].append(name)
    return players, dict(teams)

# ---------- Handicap medians ----------
def parse_hdcp_medians(wb):
    """Median handicap per player from the season's handicap_log (Name, Date, Handicap)."""
    if 'handicap_log' not in wb.sheetnames:
        return {}
    vals = defaultdict(list)
    for r in wb['handicap_log'].iter_rows(min_row=2, values_only=True):
        if not r or not r[0]: continue
        name = fix_name(str(r[0]).strip())
        if is_junk(name) or is_team_row(name): continue
        try:
            vals[name].append(float(r[2]))
        except (TypeError, ValueError):
            continue  # 'NH' / blanks
    return {name: round(statistics.median(v), 2) for name, v in vals.items() if v}

# ---------- Standings ----------
def parse_standings(wb, sheet):
    if sheet not in wb.sheetnames: return []
    rows = rows_of(wb[sheet])
    hdr = [str(clean(c) or '') for c in rows[0]]
    # find the points + pts/match columns by header text
    pts_col, ppm_col = None, None
    for i, h in enumerate(hdr):
        hl = h.lower()
        if pts_col is None and 'point' in hl and 'match' not in hl and '/' not in hl:
            pts_col = i
        if 'points/match' in hl or 'pts/match' in hl:
            ppm_col = i
    if pts_col is None: pts_col = 7
    if ppm_col is None: ppm_col = pts_col + 1
    out = []
    for r in rows[1:]:
        if len(r) <= max(pts_col, ppm_col): continue
        team, p1, p2 = num(r[0]), clean(r[1]), clean(r[2])
        if team is None or not p1: continue
        out.append({
            'team': int(team), 'p1': p1, 'p2': p2,
            'matches': num(r[3]), 'w': num(r[4]), 'l': num(r[5]), 't': num(r[6]),
            'pts': num(r[pts_col]), 'ptsMatch': num(r[ppm_col]),
        })
    return out

def parse_combined(wb, year):
    rows = rows_of(wb['season1&2'])
    out = []
    for r in rows[1:]:
        team, p1, p2 = num(r[0]), clean(r[1]), clean(r[2])
        if team is None or not p1: continue
        if year == 2025:
            rec = {'team': int(team), 'p1': p1, 'p2': p2, 'matches': num(r[3]),
                   'w': num(r[4]), 'l': num(r[5]), 't': num(r[6]),
                   'matchPts': num(r[7]), 'scramblePts': num(r[8]), 'totalPts': num(r[9]),
                   'ptsMatch': num(r[10]), 'seed': clean(r[11]), 'note': clean(r[12])}
        else:  # 2023/2024: Scramble/Medley col 7, points col 8, pts/match 9, total 10, seed 11, note 12
            rec = {'team': int(team), 'p1': p1, 'p2': p2, 'matches': num(r[3]),
                   'w': num(r[4]), 'l': num(r[5]), 't': num(r[6]),
                   'scramblePts': num(r[7]), 'matchPts': num(r[8]),
                   'ptsMatch': num(r[9]), 'totalPts': num(r[10]),
                   'seed': clean(r[11]), 'note': clean(r[12]) if len(r) > 12 else None}
        out.append(rec)
    return out

# ---------- Individual season stats ----------
def parse_individual(wb, year):
    rows = rows_of(wb['individual'])
    hdr = [clean(c) for c in rows[0]]
    has_hdcp = 'Hdcp' in hdr  # 2024 quirk
    off = 1 if has_hdcp else 0
    out = []
    for r in rows[1:]:
        name = fix_name(clean(r[0]))
        if is_junk(name) or is_team_row(name): continue
        if num(r[1]) is None: continue
        out.append({
            'name': name, 'matches': num(r[1]), 'w': num(r[2]), 'l': num(r[3]), 't': num(r[4]),
            'pts': num(r[5]), 'ptsMatch': num(r[6+off]), 'grossAvg': num(r[7+off]),
            'netAvg': num(r[8+off]), 'matchupNet': num(r[9+off]),
            'eagles': num(r[10+off]), 'birdies': num(r[11+off]), 'pars': num(r[12+off]),
            'bogeys': num(r[13+off]), 'doubles': num(r[14+off]), 'triples': num(r[15+off]),
        })
    return out

# ---------- Weekly matchups ----------
def parse_weekly(wb, year):
    rows = rows_of(wb['weekly_results'])
    matchups = []
    season = 1
    i = 0
    while i < len(rows):
        r = rows[i]
        c2 = clean(r[2]) if len(r) > 2 else None
        c0 = clean(r[0]) if len(r) > 0 else None
        if isinstance(c0, str) and c0.startswith('Season 2'): season = 2
        if isinstance(c2, str) and c2.startswith('Season 2'): season = 2
        if c2 == 'Week' and i+1 < len(rows) and num(rows[i+1][2]) is not None:
            week = int(num(rows[i+1][2]))
            course = clean(rows[i+1][7]) or ''
            j = i + 3
            while j < len(rows):
                mr = rows[j]
                if all(clean(x) is None for x in mr[:12]): break
                if clean(mr[2]) == 'Week': break
                p1, p2 = fix_name(clean(mr[0])), fix_name(clean(mr[1]))
                if is_junk(p1):
                    j += 1; continue
                def val(x):
                    v = num(x)
                    return v
                m = {'year': year, 'season': season, 'week': week, 'course': course,
                     'scramble': is_team_row(p1) or is_team_row(p2),
                     'p1': p1, 'p2': p2,
                     'p1pts': val(mr[3]), 'p1gross': val(mr[4]), 'p1net': val(mr[5]),
                     'p2pts': val(mr[9]), 'p2gross': val(mr[10]), 'p2net': val(mr[11])}
                # skip matchups with no data at all
                if m['p1pts'] is not None or m['p2pts'] is not None:
                    matchups.append(m)
                j += 1
            i = j
        else:
            i += 1
    return matchups

# ---------- Raw sheets: hole-level + dates ----------
RAW_SHEETS = {
    2023: ['gross_national','highland','columbia','columbia_front','keller','theo_wirth',
           'victory_links','victory_links_back','inver_wood','meadowbrook','les_bolstad','loggers_trail'],
    2024: ['gross_national','highland','edinburgh','edinburgh_back','inver_wood','inver_wood_back',
           'theo_wirth','columbia','columbia_back','meadowbrook','meadowbrook_back','baker_natl',
           'les_bolstad','keller','keller_back'],
    2025: ['gross_national','highland','highland_back','columbia','columbia_back','meadowbrook',
           'baker_natl','baker_natl_back','edinburgh','pioneer_creek','keller','keller_back'],
    2026: ['edinburgh','edinburgh_back','keller','keller_back','chaska_town','chaska_town_back',
           'baker_natl','baker_natl_back','pioneer_creek','keller2','keller2_back'],
}
COURSE_LABEL = {
    'gross_national':'Gross National','highland':'Highland','highland_back':'Highland',
    'columbia':'Columbia','columbia_front':'Columbia','columbia_back':'Columbia',
    'keller':'Keller','keller_back':'Keller','keller2':'Keller','keller2_back':'Keller',
    'theo_wirth':'Theodore Wirth','victory_links':'Victory Links','victory_links_back':'Victory Links',
    'inver_wood':'Inver Wood','inver_wood_back':'Inver Wood','meadowbrook':'Meadowbrook',
    'meadowbrook_back':'Meadowbrook','les_bolstad':'Les Bolstad','loggers_trail':'Loggers Trail',
    'edinburgh':'Edinburgh','edinburgh_back':'Edinburgh','baker_natl':'Baker National',
    'baker_natl_back':'Baker National','pioneer_creek':'Pioneer Creek',
    'chaska_town':'Chaska Town','chaska_town_back':'Chaska Town',
}

def parse_raw(wb, year):
    """Return list of rounds: {player, course, date, holes:[{par,score}], gross, net}"""
    rounds = []
    for sheet in RAW_SHEETS[year]:
        if sheet not in wb.sheetnames: continue
        rows = rows_of(wb[sheet])
        # find par row: col4 == 'Par' AND all values look like real pars (3-6).
        # Some sheets have 'Par' as a header label over hole numbers, so validate.
        par = None
        for r in rows[:8]:
            if clean(r[4]) == 'Par':
                cand = [num(x) for x in r[5:14]]
                if all(v is not None and 3 <= v <= 6 for v in cand):
                    par = cand
                    break
        if not par: continue
        course = COURSE_LABEL.get(sheet, sheet)
        hole_offset = 9 if 'back' in sheet else 0
        i = 0
        while i < len(rows):
            r = rows[i]
            if clean(r[0]) == 'Name' and i+2 < len(rows):
                p1 = fix_name(clean(rows[i+1][0])); d1 = datestr(rows[i+1][1]); y1 = dateyear(rows[i+1][1])
                p2 = fix_name(clean(rows[i+2][0])); d2 = datestr(rows[i+2][1]); y2 = dateyear(rows[i+2][1])
                # Workbooks can contain stale template data from prior seasons;
                # drop any round whose recorded date is from a different year.
                if y1 is not None and y1 != year: p1 = None
                if y2 is not None and y2 != year: p2 = None
                if is_junk(p1) and is_junk(p2):
                    i += 1; continue
                if is_junk(p1): p1 = None
                if is_junk(p2): p2 = None
                if is_team_row(p1): p1 = None
                if is_team_row(p2): p2 = None
                # find gross rows
                gidx = []
                for j in range(i+3, min(i+16, len(rows))):
                    if clean(rows[j][0]) == 'Gross': gidx.append(j)
                    if len(gidx) == 2: break
                for (pname, pdate, gi) in [(p1, d1, gidx[0] if len(gidx)>0 else None),
                                           (p2, d2, gidx[1] if len(gidx)>1 else None)]:
                    if not pname or gi is None: continue
                    g = [num(x) for x in rows[gi][5:14]]
                    n = [num(x) for x in rows[gi+1][5:14]] if gi+1 < len(rows) else [None]*9
                    if not any(x is not None for x in g):
                        continue
                    holes = []
                    for h in range(9):
                        if g[h] is not None and par[h] is not None:
                            holes.append({'par': int(par[h]), 'score': g[h]})
                        else:
                            holes.append(None)
                    gross_tot = sum(x for x in g if x is not None) if all(x is not None for x in g) else None
                    net_vals = [x for x in n if x is not None]
                    net_tot = round(sum(net_vals),1) if len(net_vals) == 9 else None
                    rounds.append({'player': pname, 'course': course, 'date': pdate,
                                   'year': year, 'holes': holes, 'holeOffset': hole_offset,
                                   'gross': gross_tot, 'net': net_tot})
                i = (gidx[1]+4) if len(gidx) > 1 else i+3
            else:
                i += 1
    return rounds

# ---------- Static playoff summaries ----------
PLAYOFFS = {
    2023: {
        'champion': 'Alex Manske & Thomas Wesling',
        'runnerUp': 'Riley Altenburg & Manni Mendiratta',
        'semifinalists': ['Kris Egan & Erik Jansa', 'Andrew Wolf & Kyle Makey'],
        'rounds': [
            {'name':'Quarterfinals','matches':[
                {'w':'Anderson & Bornhorst','l':'Egan & Jansa','score':'Lost on points/tiebreak — Egan & Jansa advanced as noted','courses':'Keller','note':'1v8'},
                {'w':'Altenburg & Mendiratta','l':'Hopkins & Mugford','score':'12 up','courses':'Gross National','note':'4v5'},
                {'w':'Wolf & Makey','l':'Pfieffer & Trow','score':'2 up','courses':'Bluff Creek','note':'2v7'},
                {'w':'Manske & Wesling','l':'Leadley & Mullins','score':'2 up','courses':'Inver Wood','note':'3v6'}]},
            {'name':'Semifinals','matches':[
                {'w':'Altenburg & Mendiratta','l':'Egan & Jansa','score':'2 up','courses':'Columbia'},
                {'w':'Manske & Wesling','l':'Wolf & Makey','score':'2 up','courses':'Stonebrooke'}]},
            {'name':'Finals','matches':[
                {'w':'Manske & Wesling','l':'Altenburg & Mendiratta','score':'2 up','courses':'Keller'}]}
        ]
    },
    2024: {
        'champion': 'Jason Leadley & Austin Mullins',
        'runnerUp': 'Matt Anderson & Brennan Bornhorst',
        'semifinalists': ['Joe Zittergruen & Andrew Wolf', 'Hank Hopkins & Jack Mugford'],
        'rounds': [
            {'name':'Quarterfinals','matches':[
                {'w':'Zittergruen & Wolf','l':'Egan & Jansa','score':'6 up','courses':'Chaska Town','note':'1v8'},
                {'w':'Anderson & Bornhorst','l':'Waggoner & Johnson','score':'5 up','courses':'Hiawatha','note':'4v5'},
                {'w':'Leadley & Mullins','l':'Tanner & Roland','score':'2 up','courses':'Keller','note':'2v7'},
                {'w':'Hopkins & Mugford','l':'Wesling & Nelson','score':'Coin flip after tie','courses':'Keller','note':'3v6'}]},
            {'name':'Semifinals','matches':[
                {'w':'Anderson & Bornhorst','l':'Zittergruen & Wolf','score':'Won in extra holes','courses':'Chaska Town','note':'1v4 decided in extras'},
                {'w':'Leadley & Mullins','l':'Hopkins & Mugford','score':'2 up','courses':'Keller'}]},
            {'name':'Finals','matches':[
                {'w':'Leadley & Mullins','l':'Anderson & Bornhorst','score':'2 up','courses':'Meadowbrook'}]}
        ]
    },
    2025: {
        'champion': 'Hank Hopkins & Jack Mugford',
        'runnerUp': 'Jason Leadley & Austin Mullins',
        'semifinalists': ['Thomas Wesling & Nick Nelson', 'Carsten Archer & Shay Bratland'],
        'rounds': [
            {'name':'Quarterfinals','matches':[
                {'w':'Wesling & Nelson','l':'Newman & Lacy','score':'6 up','courses':'Theodore Wirth','note':'4v6'},
                {'w':'Leadley & Mullins','l':'Gallagher & Gallagher','score':'3 up','courses':'Gross National','note':'3v5'}]},
            {'name':'Semifinals','matches':[
                {'w':'Hopkins & Mugford','l':'Wesling & Nelson','score':'1 up','courses':'Pioneer Creek','note':'1v4'},
                {'w':'Leadley & Mullins','l':'Archer & Bratland','score':'7 up','courses':'Deer Run','note':'2v5'}]},
            {'name':'Finals','matches':[
                {'w':'Hopkins & Mugford','l':'Leadley & Mullins','score':'1 up','courses':'Loggers Trail'}]}
        ]
    },
}

# ---------- Build ----------
def main():
    data = {'years': {}, 'allTime': {}, 'profiles': {}}
    all_matchups = []
    all_rounds = []
    all_individual = defaultdict(lambda: defaultdict(dict))  # player -> year -> stats
    partner_map = defaultdict(dict)  # player -> year -> partner

    hdcp_by_year = {}  # year -> {name -> median hdcp}
    for year, path in FILES.items():
        wb = openpyxl.load_workbook(path, data_only=True)
        players, teams = parse_directory(wb)
        hdcp_by_year[year] = parse_hdcp_medians(wb)
        for t, members in teams.items():
            for m in members:
                others = [x for x in members if x != m]
                if others: partner_map[m][year] = others[0]
        s1 = parse_standings(wb, 'season1')
        s2 = parse_standings(wb, 'season2')
        combined = parse_combined(wb, year) if year != 2026 else []
        indiv = parse_individual(wb, year)
        for p in indiv:
            all_individual[p['name']][year] = p
        matchups = parse_weekly(wb, year)
        all_matchups.extend(matchups)
        rounds = parse_raw(wb, year)
        all_rounds.extend(rounds)
        courses = sorted(set(m['course'] for m in matchups if m['course']))
        data['years'][year] = {
            's1': s1, 's2': s2, 'combined': combined,
            'individual': indiv, 'matchups': matchups, 'courses': courses,
            'playoffs': PLAYOFFS.get(year),
            'inProgress': year == 2026,
        }
        print(f"{year}: {len(s1)} s1 teams, {len(indiv)} players, {len(matchups)} matchup rows, {len(rounds)} rounds, courses: {courses}")

    # ---- All-time player aggregates ----
    career = {}
    for name, yearly in all_individual.items():
        m = sum(v['matches'] or 0 for v in yearly.values())
        w = sum(v['w'] or 0 for v in yearly.values())
        l = sum(v['l'] or 0 for v in yearly.values())
        t = sum(v['t'] or 0 for v in yearly.values())
        pts = sum(v['pts'] or 0 for v in yearly.values())
        # weighted avgs by matches
        def wavg(key):
            tot, cnt = 0, 0
            for v in yearly.values():
                if v.get(key) is not None and v['matches']:
                    tot += v[key] * v['matches']; cnt += v['matches']
            return round(tot/cnt, 2) if cnt else None
        career[name] = {
            'name': name, 'seasons': sorted(yearly.keys()), 'matches': m, 'w': w, 'l': l, 't': t,
            'pts': round(pts,1), 'ptsMatch': round(pts/m, 2) if m else None,
            'grossAvg': wavg('grossAvg'), 'netAvg': wavg('netAvg'),
            'eagles': sum(v['eagles'] or 0 for v in yearly.values()),
            'birdies': sum(v['birdies'] or 0 for v in yearly.values()),
        }
    data['allTime']['players'] = sorted(career.values(), key=lambda x: -(x['ptsMatch'] or 0))

    # ---- H2H per player ----
    h2h = defaultdict(lambda: defaultdict(lambda: {'meetings':0,'w':0,'l':0,'t':0,'marginSum':0.0,'details':[]}))
    for m in all_matchups:
        if m['p1pts'] is None or m['p2pts'] is None: continue
        if m.get('scramble'): continue
        a, b = m['p1'], m['p2']
        for (me, opp, mypts, opppts, mygross, mynet) in [
                (a,b,m['p1pts'],m['p2pts'],m['p1gross'],m['p1net']),
                (b,a,m['p2pts'],m['p1pts'],m['p2gross'],m['p2net'])]:
            rec = h2h[me][opp]
            rec['meetings'] += 1
            margin = mypts - opppts
            rec['marginSum'] += margin
            if margin > 0: rec['w'] += 1
            elif margin < 0: rec['l'] += 1
            else: rec['t'] += 1
            rec['details'].append({'year': m['year'], 'week': m['week'], 'course': m['course'],
                                   'myPts': mypts, 'oppPts': opppts})

    # ---- Rounds by player for course stats + hole records + best rounds ----
    rounds_by_player = defaultdict(list)
    for r in all_rounds:
        rounds_by_player[r['player']].append(r)

    # Points per (player, year, course, week) via matchups for best-points lookup
    def best_matches(name):
        best_pts, best_gross, best_net = None, None, None
        for m in all_matchups:
            if m.get('scramble'): continue
            for (me, opp, pts, gross, net) in [(m['p1'],m['p2'],m['p1pts'],m['p1gross'],m['p1net']),
                                               (m['p2'],m['p1'],m['p2pts'],m['p2gross'],m['p2net'])]:
                if me != name: continue
                ctx = {'year': m['year'], 'week': m['week'], 'course': m['course'], 'opp': opp}
                if pts is not None and (best_pts is None or pts > best_pts['value']):
                    best_pts = {'value': pts, **ctx}
                if gross is not None and (best_gross is None or gross < best_gross['value']):
                    best_gross = {'value': gross, **ctx}
                if net is not None and (best_net is None or net < best_net['value']):
                    best_net = {'value': net, **ctx}
        return best_pts, best_gross, best_net

    # course avg pts per player
    course_pts = defaultdict(lambda: defaultdict(list))
    for m in all_matchups:
        if m.get('scramble'): continue
        for (me, pts) in [(m['p1'], m['p1pts']), (m['p2'], m['p2pts'])]:
            if pts is not None and m['course']:
                course_pts[me][m['course']].append(pts)

    VSPAR_NAME = {-3:'Albatross',-2:'Eagle',-1:'Birdie',0:'Par',1:'Bogey',2:'Double Bogey',3:'Triple Bogey'}
    def vspar_label(d):
        return VSPAR_NAME.get(d, f'+{d}' if d > 0 else str(d))

    # ---- Build profiles ----
    all_players = sorted(n for n in set(list(career.keys()) + list(rounds_by_player.keys()))
                         if not is_junk(n) and not is_team_row(n))
    for name in all_players:
        prof = {'name': name}
        prof['partners'] = partner_map.get(name, {})
        # Handicap: per-season medians + average across seasons played (static years only;
        # client adds 2026 live). Stored raw so the client can merge cleanly.
        season_medians = {}
        for yr, mp in hdcp_by_year.items():
            if name in mp: season_medians[str(yr)] = mp[name]
        prof['hdcpMedians'] = season_medians
        if career.get(name):
            career[name]['avgHdcp'] = (round(sum(season_medians.values())/len(season_medians), 2)
                                       if season_medians else None)
        prof['career'] = career.get(name)
        prof['yearly'] = {str(y): v for y, v in sorted(all_individual.get(name, {}).items())}

        # H2H base: raw per-opponent tallies; client merges live 2026 and computes
        # easiest/hardest so current-season results are always reflected.
        opps = h2h.get(name, {})
        prof['h2hBase'] = {opp: {'meetings': rec['meetings'], 'w': rec['w'], 'l': rec['l'],
                                 't': rec['t'], 'marginSum': round(rec['marginSum'], 2)}
                           for opp, rec in opps.items()}
        # Course base: raw sums from matchup rows (pts + gross + net); client merges 2026.
        course_base = {}
        for m in all_matchups:
            if m.get('scramble'): continue
            for (me, pts, gross, net_) in [(m['p1'], m['p1pts'], m['p1gross'], m['p1net']),
                                           (m['p2'], m['p2pts'], m['p2gross'], m['p2net'])]:
                if me != name or pts is None or not m['course']: continue
                cb = course_base.setdefault(m['course'], {'rounds':0,'ptsSum':0,'grossSum':0,'grossN':0,'netSum':0,'netN':0})
                cb['rounds'] += 1
                cb['ptsSum'] += pts
                if gross is not None: cb['grossSum'] += gross; cb['grossN'] += 1
                if net_ is not None: cb['netSum'] += net_; cb['netN'] += 1
        for cb in course_base.values():
            cb['ptsSum'] = round(cb['ptsSum'], 1); cb['netSum'] = round(cb['netSum'], 1)
        prof['courseBase'] = course_base

        # best match performances
        bp, bg, bn = best_matches(name)
        prof['bestPts'], prof['bestGross'], prof['bestNet'] = bp, bg, bn
        # attach dates from rounds where possible (match by year+course+gross)
        def find_date(best, key):
            if not best: return
            for r in rounds_by_player.get(name, []):
                if r['year'] == best['year'] and r['course'] == best['course'] and r.get(key) == best['value']:
                    best['date'] = r['date']; return
        find_date(bg, 'gross'); find_date(bn, 'net')
        if bp:
            for r in rounds_by_player.get(name, []):
                if r['year'] == bp['year'] and r['course'] == bp['course']:
                    bp['date'] = r['date']; break

        # hole records
        best_hole, worst_hole = None, None
        for r in rounds_by_player.get(name, []):
            for hi, h in enumerate(r['holes']):
                if not h: continue
                d = h['score'] - h['par']
                rec = {'hole': hi+1+r.get('holeOffset', 0), 'par': h['par'], 'score': int(h['score']) if h['score'] == int(h['score']) else h['score'],
                       'vsPar': int(d) if d == int(d) else d, 'label': vspar_label(int(d)) if d == int(d) else None,
                       'course': r['course'], 'year': r['year'], 'date': r['date']}
                if best_hole is None or d < best_hole['vsPar']: best_hole = rec
                if worst_hole is None or d > worst_hole['vsPar']: worst_hole = rec
        prof['bestHole'], prof['worstHole'] = best_hole, worst_hole

        # fun fact candidates
        facts = []
        c = career.get(name)
        if c:
            if c['t'] and c['matches'] and c['t'] / c['matches'] >= 0.25:
                facts.append(f"Ties {round(100*c['t']/c['matches'])}% of matches — the league's great equalizer.")
            if c['eagles'] and c['eagles'] >= 2:
                facts.append(f"Has {c['eagles']} career eagles — tied for flexing rights.")
            elif c['eagles'] == 1:
                facts.append("Has a career eagle on the books.")
        # win streak
        my_matches = sorted([mm for mm in all_matchups if name in (mm['p1'], mm['p2'])
                             and not mm.get('scramble')
                             and mm['p1pts'] is not None and mm['p2pts'] is not None],
                            key=lambda m: (m['year'], m['season'], m['week']))
        streak, best_streak = 0, 0
        for mm in my_matches:
            mine = mm['p1pts'] if mm['p1'] == name else mm['p2pts']
            theirs = mm['p2pts'] if mm['p1'] == name else mm['p1pts']
            if mine > theirs: streak += 1; best_streak = max(best_streak, streak)
            else: streak = 0
        if best_streak >= 3:
            facts.append(f"Longest match win streak: {best_streak} in a row.")
        # undefeated vs someone with 3+ meetings
        for opp, rec in opps.items():
            if rec['meetings'] >= 3 and rec['l'] == 0:
                facts.append(f"Has never lost to {opp} in {rec['meetings']} meetings.")
                break
        # 10-point sweep
        for mm in my_matches:
            mine = mm['p1pts'] if mm['p1'] == name else mm['p2pts']
            if mine == 10:
                facts.append(f"Once swept a match 10–0 ({mm['course']}, {mm['year']}).")
                break
        prof['funFact'] = facts[0] if facts else None
        prof['funFacts'] = facts[:3]

        data['profiles'][name] = prof

    # h2h rivalry grid is computed client-side from profiles' h2hBase + live 2026

    with open('/home/claude/history.json', 'w') as f:
        json.dump(data, f, separators=(',', ':'))
    import os
    print(f"\nWrote history.json ({os.path.getsize('/home/claude/history.json')//1024} KB)")
    print(f"Players: {len(data['profiles'])}, matchups: {len(all_matchups)}, rounds: {len(all_rounds)}")

main()
